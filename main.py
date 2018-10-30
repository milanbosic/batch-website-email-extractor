import fnmatch
import os
import re

import openpyxl
import validators
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchFrameException

# Array to hold all emails
emails = []
urls = []

# Automatically find the excel file in the current directory
excel_file = None
for file in os.listdir('.'):
    if fnmatch.fnmatch(file, '*.xlsx'):
        excel_file = file
        print('Opening file: ' + '\'' + file.title() + '\'')
if excel_file is not None:
    wb = openpyxl.load_workbook(excel_file)
else:
    exit('No excel file found')

# First sheet
ws = wb[wb.sheetnames[0]]

# Selenium Chrome Webdriver
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('windows-size=1200x600')
chrome_options.add_argument('--disable-gpu')

driver = webdriver.Chrome('chromedriver', options=chrome_options)

# Starting parameters
column_with_urls = input('Enter the letter of the column with urls: ')
column_emails = input('Enter the letter of the column to write emails to: ')
starting_row = int(input('Enter the starting row: '))
number_of_rows = int(input('Enter the number of rows to process: '))

keywords_to_avoid = ['@example', 'example@', 'broofa', '@sentry', 'yourcompany@', 'godaddy', 'placeholder', 'name@',
                     '@domain', '.png', '.jpg', '.gif', '.jpeg']


def get_valid_url(base_url, curr_url):

    if validators.url(curr_url):
        return curr_url
    else:
        if (curr_url.startswith('/') and not base_url.endswith('/')) or (
                not curr_url.startswith('/') and base_url.endswith('/')):
            concatenated_url = base_url + curr_url
        elif curr_url.startswith('/') and base_url.endswith('/'):
            concatenated_url = base_url + curr_url[1:]
        else:
            concatenated_url = base_url + '/' + curr_url

    if validators.url(concatenated_url):
        return concatenated_url
    return None


# Checks for emails in the page source
def scrape_emails(curr_driver, search_for_links):
    # test_text = 'asdf@asdf.com example@asdf.com milan@gamil.com milan@asdf.jpg'
    match = re.findall(r"([\w\.-]+@[\w\.-]+\.\w+)", curr_driver.page_source)
    for email in match:
        if not any(x in email.lower() for x in keywords_to_avoid):
            if email not in emails:
                emails.append(email)
                print('Found email: ' + email)
    if search_for_links:
        check_for_links(curr_driver)


def check_for_links(curr_driver):

    for a in curr_driver.find_elements_by_xpath("//a[@href]"):
        temp = a.get_attribute('text') + a.get_attribute('href')
        if any(x in temp.lower() for x in ['contact', 'connect', 'location', 'about', 'welcome', 'support']):
            link = a.get_attribute('href')
            valid_url = get_valid_url(url, link)
            if valid_url and len(urls) < 10:
                urls.append(valid_url)


def check_frames_for_emails(curr_driver, search_for_links):
    # Iterate through a list of frames of type 'iframe' or 'frame'
    # and check for emails and relevant links
    frames = curr_driver.find_elements_by_tag_name('iframe')
    # frames.append(curr_driver.find_elements_by_tag_name('frame'))
    for frame in frames:
        try:
            curr_driver.switch_to.frame(frame)
        except NoSuchFrameException:
            pass
        scrape_emails(curr_driver, False)
        if search_for_links:
            check_for_links(curr_driver)

        curr_driver.switch_to.default_content()

    frames = curr_driver.find_elements_by_tag_name('frame')
    for frame in frames:
        try:
            curr_driver.switch_to.frame(frame)
        except NoSuchFrameException:
            pass
        scrape_emails(curr_driver, False)
        if search_for_links:
            check_for_links(curr_driver)

        curr_driver.switch_to.default_content()

        # TODO: Figure out how to handle frames within frames
        # if curr_driver.find_elements_by_tag_name('iframe').count() > 0 or \
        #         curr_driver.find_elements_by_tag_name('frame').count() > 0:
        #     check_frame_for_emails(curr_driver)


def get_most_relevant_email(emails):
    for email in emails:
        if any(x in email for x in ['info', 'contact']):
            return email
    return emails[0]


# Iterate through the rows with urls
for row_number in range(starting_row, starting_row + number_of_rows):
    try:
        url = ws[column_with_urls + str(row_number)].value
    except Exception:
        exit('Failed to access cell at ' + column_with_urls + str(row_number) + ', exiting...')
    emails.clear()
    urls.clear()
    # Skip the row if it's a duplicate
    last_url = ws[column_with_urls + str(row_number - 1)]
    if last_url != url:
        print('Row: ' + str(row_number) + ' , URL: ' + url)
        driver.get(url)
        source_html = driver.page_source

        # First check base html for emails
        scrape_emails(driver, True)
        # If no emails have been found, check within frames
        if not emails:
            check_frames_for_emails(driver, True)

        if source_html is not None:
            if len(source_html) < 100:
                emails.append('dead URL')
                print('Probably invalid HTML')

        if not emails:
            for url in urls:
                driver.get(url)
                scrape_emails(driver, False)
                check_frames_for_emails(driver, False)

        if not emails:
            emails.append('no')

    chosen_email = get_most_relevant_email(emails)
    print('Chosen email: ' + chosen_email)
    ws[column_emails + str(row_number)] = chosen_email
    wb.save(excel_file.title())
